import openpyxl
import pandas as pd

def create_matlab_input_file(serial_number, filespec, num_cal_points, full_scale):
    # header = ['serialno', 'full_scale_flow', 'set_flow', 'raw_data']
    csv_file = filespec + '__0.csv'
    xlsx_file = filespec + '__0.xlsx'
    csv_fh = pd.read_csv(csv_file)   # Reading the csv file
    xlsx_fh = pd.ExcelWriter(xlsx_file, engine='openpyxl')
    csv_fh.to_excel(xlsx_fh, index=False)
    xlsx_fh.save()

    if serial_number == '':
        serial_number = "Forgot0123456789"

    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb.active
    ws.delete_cols(idx=1, amount=3)
    ws.delete_cols(idx=3, amount=4)
    ws.insert_cols(idx=1, amount=2)
    ws['A1'] = 'serialno'               # Create the headers
    ws['B1'] = 'full_scale_flow'
    ws['C1'] = 'set_flow'
    ws['D1'] = 'raw_data'
    num_rows = ws.max_row

    ws.delete_rows(idx=num_cal_points+2, amount=(num_rows-(num_cal_points+1)))
    for index in range(0, (num_cal_points)):
        cell_A = 'A%d'  % (index + 2)
        ws[cell_A] = serial_number
        cell_B = 'B%d'  % (index + 2)
        ws[cell_B] = full_scale

    print(num_rows)
    wb.save(xlsx_file)      # Save Excel file
    wb.close()

    xlsx_fh = pd.read_excel(xlsx_file)
    result = xlsx_fh.sort_values(by=['set_flow'], ascending=True)
    result.to_excel(xlsx_file, index=False)
    print(result)

if __name__ == '__main__':
    filespec = "Data files\\Char Test Data\\Liquid Flow Char Test Data 15 Mar 2022 11_29_17"
    serial_number = ""
    create_matlab_input_file(serial_number, filespec, 17, 300)
